import dash
from dash import dcc
from dash import html
import pandas as pd
import sqlite3
import openpyxl
from statistics import mean
from dash_iconify import DashIconify
from dash.dependencies import Input, Output


con = sqlite3.connect('pythonDB.db')
def export_to_sqlite():
    '''Экспорт данных из xlsx в sqlite'''

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    #prj_dir = os.path.abspath(os.path.curdir)

    #a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # Имя базы
    #base_name = 'vladimirov_data.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = con.cursor()

    # создание таблицы если ее не существует
    cursor.execute('''CREATE TABLE IF NOT EXISTS all_data (Order_number TEXT, 
                    Order_type TEXT,
                    Storage_period TEXT,
                    Status TEXT,
                    Expected_date_of_receipt TEXT,
                    Actual_receipt_date TEXT,
                    Date_of_issue TEXT,
                    Payment_type TEXT,
                    Order_price REAL,
                    Number_of_places INT,
                    Places TEXT)''')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('all_data.xlsx', data_only=True)
    sheet = file_to_read.active

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 12):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять   
            data.append(value)
            # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        #cursor.execute("INSERT INTO all_data VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10]))
    # сохраняем изменения
    con.commit()
    # закрытие соединения
    #con.close()
# Средняя стоимость заказов по дням
df1 = pd.read_sql(
    (f'''SELECT strftime("%Y-%m-%d", Date_of_issue) as dateavg, 
    ROUND(AVG(Order_price), 2) as priceavg 
    FROM all_data GROUP BY dateavg;'''), con)
# Среднее количество заказов в день
df1_1 = pd.read_sql(
    (f'''SELECT strftime("%Y-%m-%d", Date_of_issue) as dateavg, 
    COUNT(Order_price) as daycountavg 
    FROM all_data GROUP BY dateavg;'''), con)



# Средняя стоимость заказов по неделям
df2 = pd.read_sql(
    (f'''SELECT strftime("%Y-%W 0", Date_of_issue) as weekdateavg, 
    ROUND(AVG(Order_price), 2) as weekpriceavg 
    FROM all_data GROUP BY weekdateavg;'''), con)
# Среднее количество заказов в неделю
df2_1 = pd.read_sql(
    (f'''SELECT strftime("%Y-%W 0", Date_of_issue) as dateavg, 
    COUNT(Order_price) as weekcounteavg 
    FROM all_data GROUP BY dateavg;'''), con)

# Средняя стоимость заказов в месяц
df3 = pd.read_sql(
    (f'''SELECT strftime("%Y-%m", Date_of_issue) as dateavg,
    ROUND(AVG(Order_price), 2) as priceavg 
    FROM all_data GROUP BY dateavg;'''), con)
# Среднее количество заказов в месяц
df3_1 = pd.read_sql(
    (f'''SELECT strftime("%Y-%m", Date_of_issue) as dateavg, 
    COUNT(Order_price) as monthcounteavg 
    FROM all_data GROUP BY dateavg;'''), con)

df_select_date = pd.read_sql(
    (f'''SELECT strftime("%Y-%m-%d", Date_of_issue) as dateavg, 
    ROUND(AVG(Order_price), 2) as priceavg, COUNT(Order_price) as daycountavg
    FROM all_data GROUP BY dateavg;'''), con)
df_select_date['Date'] = pd.to_datetime(df_select_date["dateavg"], format="%Y-%m-%d")
df_select_date.sort_values("Date", inplace=True)


df_amount7500 = pd.read_sql(
    (f'''SELECT COUNT(Order_number) as amount FROM all_data WHERE Order_price > 7500;'''), con)

df_amount7500_choose = pd.read_sql(
    (f'''SELECT strftime("%Y-%m-%d 00:00:00", Date_of_issue) as dateavg,
    COUNT(Order_number) as amount
    FROM all_data  WHERE Order_price > 7500 GROUP BY dateavg;'''), con)
df_amount7500_choose['Date'] = pd.to_datetime(df_amount7500_choose["dateavg"], format="%Y-%m-%d 00:00:00")
df_amount7500_choose.sort_values("Date", inplace=True)


daily_price_storage = pd.read_sql(
    (f'''SELECT ROUND(AVG(pricedailysum), 1) as pricedailysumavg, ROUND(AVG(amountdailysum), 1) as amountdailysumavg FROM(
            SELECT strftime("%Y-%m-%d", Date_of_issue) as dateavg, 
            SUM(Order_price) as pricedailysum, COUNT(Order_price) as amountdailysum
            FROM all_data WHERE dateavg IS NOT NULL GROUP BY dateavg);'''), con)

weekly_price_storage = pd.read_sql(
    (f'''SELECT ROUND(AVG(pricedailysum), 1) as priceweeklysumavg, ROUND(AVG(amountdailysum), 1) as amountweeklysumavg FROM(
            SELECT strftime("%Y-%W 0", Date_of_issue) as dateavg, 
            SUM(Order_price) as pricedailysum, COUNT(Order_price) as amountdailysum
            FROM all_data WHERE dateavg IS NOT NULL GROUP BY dateavg);'''), con)

monthly_price_storage = pd.read_sql(
    (f'''SELECT ROUND(AVG(pricedailysum), 1) as pricemonthlysumavg, ROUND(AVG(amountdailysum), 1) as amountmonthlysumavg FROM(
            SELECT strftime("%Y-%m", Date_of_issue) as dateavg, 
            SUM(Order_price) as pricedailysum, COUNT(Order_price) as amountdailysum
            FROM all_data GROUP BY dateavg);'''), con)

select_price_storage = pd.read_sql(
    (f'''SELECT strftime("%Y-%m-%d 00:00:00", Date_of_issue) as dateavg, 
    SUM(Order_price) as priceavg, COUNT(Order_price) as daycountavg
    FROM all_data GROUP BY dateavg;'''), con)
select_price_storage['Date'] = pd.to_datetime(df_select_date["dateavg"], format="%Y-%m-%d")
select_price_storage.sort_values("Date", inplace=True)
    
    
external_stylesheets = [
    {        
        "href": "https://fonts.googleapis.com/css2?",
        "family": "Lato:wght@400;700&display=swap",
        "rel": "stylesheet",
    },
]
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)
app.title = "Аналитика данных вашего ПВЗ"
colors = {
    'background': '#fff',
    'text': '#9e9e9e'
}


tabs_styles = {
    "flex-direction": "row",
    "textAlign": 'top',
    "height": '40px',
    
}

tab_style = {
    "padding": "10px",
    "color": '#AEAEAE',
    "fontSize": '12px',
    "width": '100px',
    "margin-left": '10px',
    "textAlign": 'center',
    "border-radius": '5px',
    "backgroundColor": '#F7F7F7',
    'border': '2px #e9e9e9 solid',
    "display": 'flex',
    "align-items": 'center',
    "justify-content": 'center'
}

tab_selected_style = {
    "padding": "10px",
    "color": '#AEAEAE',
    "width": '100px',
    "margin-left": '10px',
    "fontSize": '12px',
    "border-radius": '5px',
    "backgroundColor": '#ffffff',
    'border-top': '2px #d6d6d6 solid',
    'border-left': '2px #d6d6d6 solid',
    'border-right': '2px #d6d6d6 solid',
    "display": 'flex',
    "align-items": 'center',
    "justify-content": 'center',
    "transition": 'all 300ms ease'
}


app.layout = html.Div(children=[
        html.Div(
        children=[
            html.Div(
                children=[
                    html.H1(children='Анализ показателей ВПЗ',
                            style={
                                'textAlign': 'center'
                            }, className="header"),]),
            html.Div(className='for_numric_data',
                children = [
                    # ---------- Первый контейнер с данными ----------
                    html.Div(className='data-numeric-container',
                        children = [
                            html.Div(className='icon-number',
                                children = [
                                    DashIconify(
                                        icon="carbon:align-box-bottom-center",
                                        width=40,
                                        className='icon',
                                    ),
                                    html.Div(className='data-numeric-info',
                                        children = [
                                            html.H4('ЗАКАЗОВ ДОРОЖЕ 7500р', className='box-text'),
                                            html.Small('выберете даты', className='muted-text')
                                        ]
                                    ),
                                    DashIconify(
                                        icon="material-symbols:keyboard-arrow-down-rounded",
                                        width=30,
                                        color='#d6d6d6',
                                        className='icon-arrow',
                                    )
                                ]
                            ),
                            html.Div(className='right-data-numeric',
                                children = [
                                    html.Details(
                                        children=[
                                            html.Summary(className='icon-on-graph',
                                                children=[
                                                    
                                                ]
                                            ),
                                            html.Div(className="choose-date",
                                                children=[
                                                    dcc.DatePickerRange(
                                                        id="date-range-7500", className='date',
                                                        day_size=39,
                                                        min_date_allowed=df_select_date.Date.min().date(),
                                                        max_date_allowed=df_select_date.Date.max().date(),
                                                        start_date=df_select_date.Date.min().date(),
                                                        end_date=df_select_date.Date.max().date(),
                                                    ),
                                                ] 
                                            )
                                        ]
                                    ),
                                    html.P(id='amount_7500')
                                ]
                            )
                        ]
                    ),
                    # ---------- Второй контейнер с данными ----------
                    html.Div(className='data-numeric-container',
                        children = [
                            html.Div(className='icon-number',
                                children = [
                                    DashIconify(
                                        icon="carbon:align-box-bottom-center",
                                        width=40,
                                        className='icon',
                                    ),
                                    html.Div(className='data-numeric-info',
                                        children = [
                                            html.H4('ВЫПЛАТА ОТ ЯНДЕКСА', className='box-text'),
                                            html.Small('выберете даты', className='muted-text')
                                        ]
                                    ),
                                    DashIconify(
                                        icon="material-symbols:keyboard-arrow-down-rounded",
                                        width=30,
                                        color='#d6d6d6',
                                        className='icon-arrow',
                                    )
                                ]
                            ),
                            html.Div(className='right-data-numeric',
                                children = [
                                    html.Details(
                                        children=[
                                            html.Summary(className='icon-on-graph',
                                                children=[
                                                    
                                                ]
                                            ),
                                            html.Div(className="choose-date",
                                                children=[
                                                    dcc.DatePickerRange(
                                                        id="date-yandex_payment", className='date',
                                                        day_size=39,
                                                        min_date_allowed=df_select_date.Date.min().date(),
                                                        max_date_allowed=df_select_date.Date.max().date(),
                                                        start_date=df_select_date.Date.min().date(),
                                                        end_date=df_select_date.Date.max().date(),
                                                    ),
                                                ] 
                                            )
                                        ]
                                    ),
                                    html.P(id='yandex_payment')
                                ]
                            )
                        ]
                    ),
                    # ---------- Третий контейнер с данными ----------
                    html.Div(className='data-numeric-container',
                        children = [
                            html.Div(className='icon-number',
                                children = [
                                    DashIconify(
                                        icon="carbon:align-box-bottom-center",
                                        width=40,
                                        className='icon',
                                    ),
                                    html.Div(className='data-numeric-info',
                                        children = [
                                            html.H4('УНИКАЛЬНЫХ КЛИЕНТОВ', className='box-text'),
                                            html.Small('выберете даты', className='muted-text')
                                        ]
                                    ),
                                    DashIconify(
                                        icon="material-symbols:keyboard-arrow-down-rounded",
                                        width=30,
                                        color='#d6d6d6',
                                        className='icon-arrow',
                                    )
                                ]
                            ),
                            html.Div(className='right-data-numeric',
                                children = [
                                    html.Details(
                                        children=[
                                            html.Summary(className='icon-on-graph',
                                                children=[
                                                    
                                                ]
                                            ),
                                            html.Div(className="choose-date",
                                                children=[
                                                    dcc.DatePickerRange(
                                                        id="date-uniq_customers", className='date',
                                                        day_size=39,
                                                        min_date_allowed=df_select_date.Date.min().date(),
                                                        max_date_allowed=df_select_date.Date.max().date(),
                                                        start_date=df_select_date.Date.min().date(),
                                                        end_date=df_select_date.Date.max().date(),
                                                    ),
                                                ] 
                                            )
                                        ]
                                    ),
                                    html.P(id='uniq_customers')
                                ]
                            )
                        ]
                    ),
                    # ---------- Четвертый контейнер с данными ----------
                    html.Div(className='data-numeric-container',
                        children = [
                            html.Div(className='icon-number',
                                children = [
                                    DashIconify(
                                        icon="carbon:align-box-bottom-center",
                                        width=40,
                                        className='icon',
                                    ),
                                    html.Div(className='data-numeric-info',
                                        children = [
                                            html.H4('ПОСТОЯННЫХ КЛИЕНТОВ', className='box-text'),
                                            html.Small('выберете даты', className='muted-text')
                                        ]
                                    ),
                                    DashIconify(
                                        icon="material-symbols:keyboard-arrow-down-rounded",
                                        width=30,
                                        color='#d6d6d6',
                                        className='icon-arrow',
                                    )
                                ]
                            ),
                            html.Div(className='right-data-numeric',
                                children = [
                                    html.Details(
                                        children=[
                                            html.Summary(className='icon-on-graph',
                                                children=[
                                                    
                                                ]
                                            ),
                                            html.Div(className="choose-date",
                                                children=[
                                                    dcc.DatePickerRange(
                                                        id="date-regular_customers", className='date',
                                                        day_size=39,
                                                        min_date_allowed=df_select_date.Date.min().date(),
                                                        max_date_allowed=df_select_date.Date.max().date(),
                                                        start_date=df_select_date.Date.min().date(),
                                                        end_date=df_select_date.Date.max().date(),
                                                    ),
                                                ] 
                                            )
                                        ]
                                    ),
                                    html.P(id='regular_customers')
                                ]
                            )
                        ]
                    ),
                ]
            ),
            html.Div(children=[
                # ----------- первый график -----------
                html.Div([
                    dcc.Tabs(id="tabs", value='tab-1', className='TabGroup', children=[
                        dcc.Tab(label='Цена', value='tab-1', style = tab_style,
                        selected_style = tab_selected_style, className='One-Tab'),
                        dcc.Tab(label='Кол-во', value='tab-12', style = tab_style,
                        selected_style = tab_selected_style,className='One-Tab'),
                    ], style = tabs_styles),
                    html.Div(id='tabs-content', className='graf-only'),
                    # --------------- Числовые данные внизу графика ---------------
                    html.Div(children = [
                            # --------------- Первая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-currency-ruble",
                                            width=20,
                                            )
                                    ], className='icon-price'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ₽'),
                                            html.Small('продано за 1 день', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(daily_price_storage['pricedailysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                            # --------------- Вторая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-production-quantity-limits",
                                            width=20,
                                            )
                                    ], className='icon-amount'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ШТ'),
                                            html.Small('продано за 1 день', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(daily_price_storage['amountdailysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                    ],
                             className='data-on-graph')
                    ], className='container-for-graph'
                ),
                # ----------- второй график -----------
                html.Div([
                    dcc.Tabs(id="tabs_2", value='tab-2', className='TabGroup', children=[
                        dcc.Tab(label='Цена', value='tab-2', style = tab_style,
                        selected_style = tab_selected_style, className='One-Tab'),
                        dcc.Tab(label='Кол-во', value='tab-22', style = tab_style,
                        selected_style = tab_selected_style, className='One-Tab'),
                    ], style = tabs_styles),
                    html.Div(id='tabs-content_2', className='graf-only'),
                    # --------------- Числовые данные внизу графика ---------------
                    html.Div(children = [
                            # --------------- Первая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-currency-ruble",
                                            width=20,
                                            )
                                    ], className='icon-price'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ₽'),
                                            html.Small('продано за 1 неделю', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(weekly_price_storage['priceweeklysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                            # --------------- Вторая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-production-quantity-limits",
                                            width=20,
                                            )
                                    ], className='icon-amount'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ШТ'),
                                            html.Small('продано за 1 неделю', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(weekly_price_storage['amountweeklysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                    ],
                             className='data-on-graph')
                    ], className='container-for-graph'
                ),
                # ----------- третий график -----------
                html.Div([
                    dcc.Tabs(id="tabs_3", value='tab-3', className='TabGroup', children=[
                        dcc.Tab(label='Цена', value='tab-3', style = tab_style,
                        selected_style = tab_selected_style, className='One-Tab'),
                        dcc.Tab(label='Кол-во', value='tab-32', style = tab_style,
                        selected_style = tab_selected_style,className='One-Tab'),
                    ], style = tabs_styles),
                    html.Div(id='tabs-content_3', className='graf-only'),
                    # --------------- Числовые данные внизу графика ---------------
                    html.Div(children = [
                            # --------------- Первая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-currency-ruble",
                                            width=20,
                                            )
                                    ], className='icon-price'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ₽'),
                                            html.Small('продано за 1 месяц', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(monthly_price_storage['pricemonthlysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                            # --------------- Вторая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-production-quantity-limits",
                                            width=20,
                                            )
                                    ], className='icon-amount'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ШТ'),
                                            html.Small('продано за 1 месяц', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(monthly_price_storage['amountmonthlysumavg'])], className='number')
                            ], className='data-on-graph-container-test'),
                    ],
                             className='data-on-graph')
                    ], className='container-for-graph'
                ),
                # ----------- четвертый график -----------
                html.Div([
                    html.Div(
                        children=[
                            html.Details(
                                children=[
                                    html.Summary(children=[
                                        DashIconify(
                                        icon="material-symbols:keyboard-arrow-down-rounded",
                                        width=30,
                                        color='#d6d6d6'
                                        )
                                    ], className='icon-on-graph'),
                                    html.Div(
                                        children=[
                                            dcc.DatePickerRange(
                                                id="date-range", className='date',
                                                day_size=39,
                                                style={'backgroundColor': '#000000'},
                                                min_date_allowed=df_select_date.Date.min().date(),
                                                max_date_allowed=df_select_date.Date.max().date(),
                                                start_date=df_select_date.Date.min().date(),
                                                end_date=df_select_date.Date.max().date(),
                                                    ),
                                        ], className="choose-date"
                                    )
                                ]
                            ),
                            html.Div(
                                children=[
                                    dcc.Tabs(id="tabs_4", value='tab-4', className='TabGroup-2',
                                            children=[
                                                dcc.Tab(label='Цена', value='tab-4', style = tab_style,
                                                selected_style = tab_selected_style, className='One-Tab',),
                                                dcc.Tab(label='Кол-во', value='tab-42', style = tab_style,
                                                selected_style = tab_selected_style,className='One-Tab'),
                                            ], style = tabs_styles
                                    )
                                ]
                            ),
                        ], className='TabGroup-3'
                    ),
                    html.Div(id='tabs-content_4', className='graf-only'),
                    # --------------- Числовые данные внизу графика ---------------
                    html.Div(children = [
                            # --------------- Первая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(children = [
                                    # --------------- Иконка ---------------
                                    html.Div(children = [
                                        DashIconify(
                                            icon="ic:sharp-currency-ruble",
                                            width=20,
                                            )
                                    ], className='icon-price'),
                                    # --------------- Только текст ---------------
                                    html.Div(children = [
                                            html.H5('СРЕДНИЕ ПРОДАЖИ, ₽'),
                                            html.Small('продано за период', className='muted-text')
                                        ], className='data-on-graph-info-test')], className='icon-and-text'),
                                # --------------- Весь текст с числом ---------------
                                html.Div(children = [
                                    html.P(id='price_data')], className='number')
                            ], className='data-on-graph-container-test'),
                            # --------------- Вторая плашка с данными ---------------
                            html.Div(children = [
                                # --------------- Иконка + текст ---------------
                                html.Div(className='icon-and-text',
                                    children = [
                                        # --------------- Иконка ---------------
                                        html.Div(className='icon-amount', children = [
                                            DashIconify(
                                                icon="ic:sharp-production-quantity-limits",
                                                width=20,
                                                )
                                        ]),
                                        # --------------- Только текст ---------------
                                        html.Div(className='data-on-graph-info-test', children = [
                                                html.H5('СРЕДНИЕ ПРОДАЖИ, ШТ'),
                                                html.Small('продано за период', className='muted-text')
                                        ])
                                    ]
                                ),
                                # --------------- Весь текст с числом ---------------
                                html.Div(className='number', children = [
                                    html.P(id='amount_data')]
                                )
                            ], className='data-on-graph-container-test'),
                        ], className='data-on-graph'
                    )
                    ], className='container-for-graph'
                ),
                # ----------- конец четвертого графика -----------
            ], className='main-container-for-graph'),
        ]
    ),
    html.Div(className='right',
        children=[
            html.Div(className='top',
                children=[
                    html.Div(className='profile',
                        children=[
                            html.Div(
                                children=[
                                    html.P('Добро пожаловать'),
                                    html.B('Антон')
                                ], className='info'
                            ),
                            html.Div(className='profile-photo',
                                children=[
                                    html.Img(src='assets/images/profile-2.jpg')
                                ]
                            )
                        ]
                    ),

                ]
            ),
            # ---------- END TOP -----------
            html.Div(className='sales-analytics',
                children=[
                    html.H2('Аналитика заказов'),
                    # ---------- FIRST ITEM -----------
                    html.Div(className='item online', 
                        children=[
                            html.Div(className='icon', 
                                children=[
                                    DashIconify(icon="ic:sharp-production-quantity-limits",
                                                width=20),
                                ]
                            ),
                            html.Div(className='right', 
                                children=[
                                    html.Div(className='info', 
                                        children=[
                                            html.H3('ПОСТУПЛЕНИЙ ВЧЕРА'),
                                            html.Small('За 24 часа', className='text-muted')
                                        ]
                                    ),
                                    html.H5('+8%', className='success'),
                                    html.H3('243')
                                ]
                            ),
                        ]
                    ),
                    # ---------- SECOND ITEM -----------
                    html.Div(className='item online', 
                        children=[
                            html.Div(className='icon', 
                                children=[
                                    DashIconify(icon="ic:sharp-production-quantity-limits",
                                                width=20),
                                ]
                            ),
                            html.Div(className='right', 
                                children=[
                                    html.Div(className='info', 
                                        children=[
                                            html.H3('ЗАБРАЛИ ЗАКАЗ ВЧЕРА'),
                                            html.Small('За 24 часа', className='text-muted')
                                        ]
                                    ),
                                    html.H5('+8%', className='success'),
                                    html.H3('243')
                                ]
                            ),
                        ]
                    ),
                    # ---------- THIRD ITEM -----------
                    html.Div(className='item online', 
                        children=[
                            html.Div(className='icon', 
                                children=[
                                    DashIconify(icon="ic:sharp-production-quantity-limits",
                                                width=20),
                                ]
                            ),
                            html.Div(className='right', 
                                children=[
                                    html.Div(className='info', 
                                        children=[
                                            html.H3('СДЕЛАЛ ЗАКАЗ ВЧЕРА'),
                                            html.Small('За 24 часа', className='text-muted')
                                        ]
                                    ),
                                    html.H5('+8%', className='success'),
                                    html.H3('243')
                                ]
                            ),
                        ]
                    ),

                ]
            )
            # ---------- END SALES ANALYTICS -----------
        ]
    )

], className='container')



@app.callback(Output('tabs-content', 'children'),
              Input('tabs', 'value'))
def render_content(tab):
    if tab == 'tab-1':
        return html.Div(children=[dcc.Graph(
            id='Средняя стоимость заказа в день',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': df1['dateavg'], 'y': df1['priceavg']},
                ],
                'layout': {'title': {'text': 'Средняя стоимость заказов по дням',
                                     'x': 0.09,
                                     "xanchor": "left",
                                    },
                           'colorway': ['#7380ec'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],

                           'font': {
                               'color': colors['text']
                           }}
            }, className="card"  
        ), 
        ])
    elif tab == 'tab-12':
        return html.Div([dcc.Graph(
            id='Среднее количество заказов по дням',
            config={"displayModeBar": False},
            style={'textAlign': 'left'},
            figure={
                'data': [
                    {'x': df1_1['dateavg'], 'y': df1_1['daycountavg']},
                ],
                'layout': {'title': {'text': 'Среднее количество заказов по дням',
                                     'x': 0.09,
                                     "xanchor": "left",
                                    },
                           'yaxis':dict(range=[0,100]),
                           'colorway': ['#ff7782'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'font': {
                                'textAlign': 'left',
                                'color': colors['text']
                           }},
                
            },
        ),
        ])
    
    
@app.callback(Output('tabs-content_2', 'children'),
              Input('tabs_2', 'value'))
def render_content(tab):
    if tab == 'tab-2':
        return html.Div([dcc.Graph(
            id='Средняя стоимость заказов в неделю',
            config={"displayModeBar": False},
            figure={
                
                'data': [
                    {'x': df2['weekdateavg'], 'y': df2['weekpriceavg']},
                ],
                'layout': {'title': {'text': 'Средняя стоимость заказов в неделю',
                                     'x': 0.09,
                                     'xanchor': "left",
                                    },
                           'colorway': ['#7380ec'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'textAlign':'left',
                           
                           'font': {
                               
                               'color': colors['text']
                           }},
            },
        ),
        ])
    elif tab == 'tab-22':
        return html.Div([dcc.Graph(
            id='Среднее количество заказов в неделю',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': df2_1['dateavg'], 'y': df2_1['weekcounteavg']},
                ],
                'layout': {'title': {'text': 'Среднее количество заказов в неделю',
                                     'x': 0.09,
                                     "xanchor": "left",
                                    },
                           'colorway': ['#ff7782'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'font': {
                               'color': colors['text']
                           }},
            },
        ),
        ])


@app.callback(Output('tabs-content_3', 'children'),
              Input('tabs_3', 'value'))
def render_content(tab):
    if tab == 'tab-3':
        return html.Div([dcc.Graph(
            id='Средняя стоимость заказов в месяц',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': df3['dateavg'], 'y': df3['priceavg']},
                ],
                'layout': {'title': {'text': 'Средняя стоимость заказов в месяц',
                                     'x': 0.09,
                                     'xanchor': "left",
                                    },
                           'colorway': ['#7380ec'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'textAlign':'left',
                           'font': {
                               
                               'color': colors['text']
                           }},
            },
        ),
        ])
    elif tab == 'tab-32':
        return html.Div([dcc.Graph(
            id='Среднее количество заказов в месяц',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': df3_1['dateavg'], 'y': df3_1['monthcounteavg']},
                ],
                'layout': {'title': {'text': 'Среднее количество заказов в месяц',
                                     'x': 0.09,
                                     "xanchor": "left",
                                    },
                           'colorway': ['#ff7782'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'font': {
                               'color': colors['text']
                           }},
            },
        ),
        ])


@app.callback(
    Output('tabs-content_4', 'children'),
    [
        Input('tabs_4', 'value'),
        Input("date-range", "start_date"),
        Input("date-range", "end_date"),
    ],
)
def update_charts(tab, start_date, end_date):
    mask = (
        (df_select_date.Date >= start_date)
        & (df_select_date.Date <= end_date)
    )
    filtered_data = df_select_date.loc[mask, :]
    if tab == 'tab-4':
        return html.Div([dcc.Graph(
            id='Средняя стоимость заказов в месяц',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': filtered_data["Date"], 'y': df_select_date['priceavg']},
                ],
                'layout': {'title': {'text': 'Средняя стоимость заказов',
                                     'x': 0.09,
                                     'xanchor': "left",
                                    },
                           'colorway': ['#7380ec'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'textAlign':'left',
                           'font': {
                               
                               'color': colors['text']
                           }},
            },
        ),
        ])
    elif tab == 'tab-42':
        return html.Div([dcc.Graph(
            id='Среднее количество заказов в месяц',
            config={"displayModeBar": False},
            figure={
                'data': [
                    {'x': filtered_data["Date"], 'y': df_select_date['daycountavg']},
                ],
                'layout': {'title': {'text': 'Среднее количество заказов',
                                     'x': 0.09,
                                     "xanchor": "left",
                                    },
                           'colorway': ['#ff7782'],
                           'plot_bgcolor': colors['background'],
                           'paper_bgcolor': colors['background'],
                           'font': {
                               'color': colors['text']
                           }},
            },
        ),
        ])

@app.callback(
    [Output('price_data', 'children'), Output('amount_data', 'children')],
    [
        Input("date-range", "start_date"),
        Input("date-range", "end_date"),
    ],
)
def output_data_price(start_date, end_date):
    mask = (
        (select_price_storage.Date >= start_date)
        & (select_price_storage.Date <= end_date)
    )
    filtered_data = select_price_storage.loc[mask, :]
    
    choose_data = filtered_data["Date"], 
    choose_date = select_price_storage['dateavg']
    choose_count = select_price_storage['daycountavg']
    choose_price = select_price_storage['priceavg']

    count_dict = {}
    for i in range(len(choose_date)):
        count_dict[choose_date[i]] = choose_count[i]

    price_dict = {}
    for i in range(len(choose_date)):
        price_dict[choose_date[i]] = choose_price[i]
    

    list_for_count = []
    for i, v in count_dict.items():
        for s in choose_data:
            for j in s:
                if str(j)==str(i):
                    list_for_count.append(int(v))
    
    list_for_price = []
    for i, v in price_dict.items():
        for s in choose_data:
            for j in s:
                if str(j)==str(i):
                    list_for_price.append(int(v))

    avg_count = round(mean(list_for_count), 1)
    avg_price = round(mean(list_for_price), 1)

    return avg_price, avg_count

@app.callback(
    Output('amount_7500', 'children'),
    [
        Input("date-range-7500", "start_date"),
        Input("date-range-7500", "end_date"),
    ],
)
def output_data_price(start_date, end_date):
    mask = (
        (df_amount7500_choose.Date >= start_date)
        & (df_amount7500_choose.Date <= end_date)
    )
    filtered_data = df_amount7500_choose.loc[mask, :]
    
    choose_data_7500 = filtered_data["Date"], 
    choose_date_7500 = df_amount7500_choose['dateavg']
    choose_count_7500 = df_amount7500_choose['amount']


    count_dict = {}
    for i in range(len(choose_date_7500)):
        count_dict[choose_date_7500[i]] = choose_count_7500[i]
    

    list_for_count = []
    for i, v in count_dict.items():
        for s in choose_data_7500:
            for j in s:
                
                if str(j)==str(i):
                    list_for_count.append(int(v))

    sum_count = sum(list_for_count)
    return sum_count


@app.callback(
    Output('yandex_payment', 'children'),
    [
        Input("date-yandex_payment", "start_date"),
        Input("date-yandex_payment", "end_date"),
    ],
)
def output_yandex_payment(start_date, end_date):
    mask = (
        (df_amount7500_choose.Date >= start_date)
        & (df_amount7500_choose.Date <= end_date)
    )
    filtered_data = df_amount7500_choose.loc[mask, :]
    
    choose_data_7500 = filtered_data["Date"], 
    choose_date_7500 = df_amount7500_choose['dateavg']
    choose_count_7500 = df_amount7500_choose['amount']


    count_dict = {}
    for i in range(len(choose_date_7500)):
        count_dict[choose_date_7500[i]] = choose_count_7500[i]
    

    list_for_count = []
    for i, v in count_dict.items():
        for s in choose_data_7500:
            for j in s:
                
                if str(j)==str(i):
                    list_for_count.append(int(v))

    sum_count = sum(list_for_count)
    return sum_count


@app.callback(
    Output('uniq_customers', 'children'),
    [
        Input("date-uniq_customers", "start_date"),
        Input("date-uniq_customers", "end_date"),
    ],
)
def output_yandex_payment(start_date, end_date):
    mask = (
        (df_amount7500_choose.Date >= start_date)
        & (df_amount7500_choose.Date <= end_date)
    )
    filtered_data = df_amount7500_choose.loc[mask, :]
    
    choose_data_7500 = filtered_data["Date"], 
    choose_date_7500 = df_amount7500_choose['dateavg']
    choose_count_7500 = df_amount7500_choose['amount']


    count_dict = {}
    for i in range(len(choose_date_7500)):
        count_dict[choose_date_7500[i]] = choose_count_7500[i]
    

    list_for_count = []
    for i, v in count_dict.items():
        for s in choose_data_7500:
            for j in s:
                
                if str(j)==str(i):
                    list_for_count.append(int(v))

    sum_count = sum(list_for_count)
    return sum_count

@app.callback(
    Output('regular_customers', 'children'),
    [
        Input("date-regular_customers", "start_date"),
        Input("date-regular_customers", "end_date"),
    ],
)
def output_yandex_payment(start_date, end_date):
    mask = (
        (df_amount7500_choose.Date >= start_date)
        & (df_amount7500_choose.Date <= end_date)
    )
    filtered_data = df_amount7500_choose.loc[mask, :]
    
    choose_data_7500 = filtered_data["Date"], 
    choose_date_7500 = df_amount7500_choose['dateavg']
    choose_count_7500 = df_amount7500_choose['amount']


    count_dict = {}
    for i in range(len(choose_date_7500)):
        count_dict[choose_date_7500[i]] = choose_count_7500[i]
    

    list_for_count = []
    for i, v in count_dict.items():
        for s in choose_data_7500:
            for j in s:
                
                if str(j)==str(i):
                    list_for_count.append(int(v))

    sum_count = sum(list_for_count)
    return sum_count
                

if __name__ == '__main__':
    app.run(debug=True)






